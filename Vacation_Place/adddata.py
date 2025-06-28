import openpyxl
from pathlib import Path
import pymongo
import dns
from datetime import datetime
from bson import Decimal128

xlsx_file = Path('./Province.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
# for row in sheet.iter_rows():
#     for cell in row:
#         print(cell[].value, end=" ")
#     print()

col_names = []

for column in sheet.iter_rows():
    col_names.append([column[0].value, column[1].value])


def addCloud():
    # mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority
    connectMongodb = 'mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority'
    with pymongo.MongoClient(connectMongodb)as conn:
        db = conn.get_database('VacationPlace')
    # for i in range(col_names.__len__()):
    #     data = {'ภาค':col_names[i][1],'Province':col_names[i][0]}
    #     currID = db.Province.insert_one(data)
    # print('well done')
    # db = conn.get_database('TNI')  # .update_one() .update_many()
    result = db.Province.update_one({'Province': 'กรุงเทพมหานคร'}, {'$set': {'attractions': [{'name': 'พิพิธภัณฑ์ชาวบางกอก',
                                                                                            'place': '273 ซอย สะพานยาว แขวง สี่พระยา เขตบางรัก กรุงเทพมหานคร 10500',
                                                                                            'openday': 'เปิดทุกวันยกเว้นวันจันทร์',
                                                                                            'opentime': 9.00,
                                                                                            'closetime': '16.00',
                                                                                            'phone': '02-233-7027',
                                                                                            'webside': '-'}
        , {'name': 'สามย่าน มิตรทาวน์',
           'place': '944 ถนนพระรามที่ ๔ เเขวง วังใหม่ เขตปทุมวัน กรุงเทพมหานคร 10330',
           'openday': 'เปิดทุกวัน',
           'opentime': 10.00,
           'closetime': '22.00',
           'phone': '02-033-8900',
           'webside': '-'},
                                                                                           # {'name': 'ธาตุก่องข้าวน้อย',
                                                                                           #  'place': 'ตำบล ตาดทอง อำเภอเมืองยโสธร ยโสธร 35000',
                                                                                           #  'openday': 'เปิดทุกวัน',
                                                                                           #  'opentime': 8.30,
                                                                                           #  'closetime': '16.30',
                                                                                           #  'phone': '-',
                                                                                           #  'webside': '-'},
                                                                                           # {
                                                                                           #     'name': 'ภูถ้ำพระ',
                                                                                           #     'place': 'ตำบล คำน้ำสร้าง อำเภอ กุดชุม ยโสธร 35140',
                                                                                           #     'openday': 'เปิดทุกวัน',
                                                                                           #     'opentime': 8.00,
                                                                                           #     'closetime': '18.00',
                                                                                           #     'phone': '-',
                                                                                           #     'webside': '-'}
                                                                                              ]}})
    print("Match :{}\nUpdate :{}".format(result.matched_count, result.modified_count))


def addrester():
    # mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority
    connectMongodb = 'mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority'
    with pymongo.MongoClient(connectMongodb)as conn:
        db = conn.get_database('VacationPlace')
    # for i in range(col_names.__len__()):
    #     data = {'ภาค':col_names[i][1],'Province':col_names[i][0]}
    #     currID = db.Province.insert_one(data)
    # print('well done')
    # db = conn.get_database('TNI')  # .update_one() .update_many()
    result = db.Province.update_one({'Province': 'ยโสธร'}, {'$set': {'restaurant': [{'name': 'ช. แม่อ้อมลาบเป็ด',
                                                                                           'place': '136 หมู่ 6 บ้านท่าเยี่ยม ต. เขื่องคำ, เมือง, ยโสธร 35000',
                                                                                           'openday': 'เปิดทุกวัน',
                                                                                           'opentime': 8.00,
                                                                                           'closetime': '16.00',
                                                                                           'phone': '083-734-2658',
                                                                                           'webside': '-'}
        , {'name': 'ร้านแซบอีหลี ยโสธร',
           'place': 'ตำบล เขื่องคำ อำเภอเมืองยโสธร ยโสธร 35000',
           'openday': 'เปิดทุกวัน',
           'opentime': 9.00,
           'closetime': '22.00',
           'phone': '085-793-2409',
           'webside': '-'},
                                                                                          {'name': 'สวนอาหารระเบียงรัก',
                                                                                           'place': 'ตำบลในเมือง อำเภอเมืองยโสธร ยโสธร 35000',
                                                                                           'openday': 'เปิดทุกวัน',
                                                                                           'opentime': 10.00,
                                                                                           'closetime': '22.00',
                                                                                           'phone': '093-118-9538',
                                                                                           'webside': '-'},
                                                                                          {
                                                                                              'name': 'บ้านคุณย่า',
                                                                                              'place': '220/10-11 ถนน รัตนเขต ตำบลในเมือง อำเภอเมืองยโสธร ยโสธร 35000',
                                                                                              'openday': 'เปิดทุกวัน',
                                                                                              'opentime': 7.30,
                                                                                              'closetime': '21.00',
                                                                                              'phone': '045-711-053',
                                                                                              'webside': '-'},
                                                                                          {
                                                                                              'name': 'ครัวราชาวดี',
                                                                                              'place': '41 13 ถนน ซอย ห้าธันวามหาราช ตำบลในเมือง อำเภอเมืองยโสธร ยโสธร 35000',
                                                                                              'openday': 'เปิดทุกวัน',
                                                                                              'opentime': 11.00,
                                                                                              'closetime': '22.00',
                                                                                              'phone': '095-612-6529',
                                                                                              'webside': '-'}]}})
    print("Match :{}\nUpdate :{}".format(result.matched_count, result.modified_count))


def addreaddhotel():
    # mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority
    connectMongodb = 'mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority'
    with pymongo.MongoClient(connectMongodb)as conn:
        db = conn.get_database('VacationPlace')
    # for i in range(col_names.__len__()):
    #     data = {'ภาค':col_names[i][1],'Province':col_names[i][0]}
    #     currID = db.Province.insert_one(data)
    # print('well done')
    # db = conn.get_database('TNI')  # .update_one() .update_many()
    result = db.Province.update_one({'Province':'ยโสธร'},
                                    {'$set': {'resting_place': [{'name': 'โรงแรมเจพี เอมเมอรัลด์',
                                                                 'place': '36 ถนน ปารา ตำบลในเมือง ยโสธร 35000',
                                                                 'openday': 'เปิดทุกวัน',
                                                                 'room': '119 room',
                                                                 'phone': '045-714-455',
                                                                 'pool': 'not have',
                                                                 'food_room':'have',
                                                                 'check_in': '14.00',
                                                                 'check_out': '12.00',
                                                                 }
                                        , {'name': 'โรงแรมอาร์พี ซิตี้ (RP City Hotel)',
                                           'place': 'ตำบลในเมือง อำเภอเมืองยโสธร ยโสธร 35000',
                                           'openday': 'เปิดทุกวัน',
                                           'room': '34 room',
                                           'phone': '045-711-648',
                                           'pool': 'have',
                                           'food_room': 'have',
                                           'check_in': '14.00',
                                           'check_out': '12.00', },
                                                                {'name': 'โรงแรมวิมานเมฆรีสอร์ทยโสธร',
                                                                 'place': '328 หมู่2 ถนนเจ้าเสด็จ, ตำบล น้ำคำใหญ่ เมือง, ยโสธร 35000',
                                                                 'openday': 'เปิดทุกวัน',
                                                                 'room': '16 room',
                                                                 'phone': '082-945-1528',
                                                                 'pool': 'not have',
                                                                 'food_room': 'not have',
                                                                 'check_in': '14.00',
                                                                 'check_out': '12.00', }]}})
                                                               # ''' {'name': 'The Avail',
                                                               #   'place': '16/70 ถนนนเรศวร ต.ประตูชัย ท่าวาสุกรี พระนครศรีอยุธยา 13000 ',
                                                               #   'openday': 'เปิดทุกวัน',
                                                               #   'room': '10 room',
                                                               #   'phone': '035-355-359',
                                                               #   'pool': 'not have',
                                                               #   'food_room': 'have',
                                                               #   'check_in': '14.00',
                                                               #   'check_out': '12.00', }'''
    print("Match :{}\nUpdate :{}".format(result.matched_count, result.modified_count))


def testCloud1():
    connectMongodb = 'mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority'
    with pymongo.MongoClient(connectMongodb)as conn:
        where = {'Province': 'กรุงเทพมหานคร'}
        db = conn.get_database('VacationPlace')
        Found = db.Province.find(where)
        # print("data in collection = {}".format(Found))
        data = []
        j = 0
        for i in Found:
            for j in range(i['attractions'].__len__()):
                print("{}".format(i['attractions'][j]['name']))


def testRemove1():
    connectMongodb = 'mongodb+srv://Kornwara:2tIbDeNWUF7Yjswi@cluster0.ezb5a.mongodb.net/<dbname>?retryWrites=true&w=majority'
    with pymongo.MongoClient(connectMongodb)as conn:
        db = conn.get_database('VacationPlace')  # .delete_one(),.delete_many()
        where = {'gpax': 3.8}
        result = db.Provice.delete_one(where)
        print("how many delete = {}".format(result.deleted_count))

addCloud()
# addrester()
# addreaddhotel()
# print(datetime.now())

# if __name__ =='__main__':
#     filename = './Province.xlsx'
#     readCSV(filename)
